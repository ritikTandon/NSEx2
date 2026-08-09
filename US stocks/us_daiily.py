import os
import datetime
import pandas as pd
import yfinance as yf
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill


# ==========================================
# HELPER FUNCTIONS
# ==========================================
def round_to_005(val):
    """Rounds a price value to the nearest 0.05 tick size."""
    if pd.isna(val) or val is None or val == 0:
        return val
    return round(float(val) * 20) / 20


def autofit_columns(worksheet):
    """Automatically adjusts column widths for columns 6 to 10."""
    for col in worksheet.iter_cols(min_col=6, max_col=9):
        max_length = 0
        column = col[0].column_letter  # Get the column letter (F, G, H, ...)

        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Add padding for visual comfort
        worksheet.column_dimensions[column].width = max_length + 2.5


# ==========================================
# CONFIGURATION
# ==========================================
shares = [
    'TSLA', 'MSTR', 'AMD', 'INTC', 'NVDA', 'QQQ', 'MU', 'GOOGL', 'SNDK',
    'MSFT', 'META', 'PLTR', 'HOOD', 'FCUV', 'SKHY', 'SPCX', 'WMT',
    'CRCL', 'COIN', 'CBRS', 'BABA', 'AMZN'
]
# shares = ['TSLA'] # Uncomment to test with just one stock first

# Target Date (Must be within the last 7 days for 1-min data)
cur_date = '07.08.26'
cur_month = 'AUG'
cur_year = 2026

OFFSET = 301 # INCREMENT THIS DAILY (daily row)

BASE_DIR = r"E:\Daily Data work\USA STOCKS"

# Styles
red = Font("Arial", 11, color='ff0000', bold=True)
blue = Font("Arial", 11, color="0000ff", bold=True)
bold = Font("Arial", 11, bold=True)
alignment = Alignment(horizontal='center')
yellow_fill = PatternFill("solid", fgColor="FFFF00")

cur_date_datetime = datetime.date(int(cur_year), int(cur_date[3:5]), int(cur_date[:2]))

# Data dictionaries
high_dict = {share: 0 for share in shares}
low_dict = {share: 0 for share in shares}
cl_9_40_dict = {share: 0 for share in shares}
prev_high_dict = {share: 0 for share in shares}
prev_low_dict = {share: 0 for share in shares}
prev_close_dict = {share: 0 for share in shares}

# ==========================================
# 1-MINUTE DATA & PREVIOUS DAY PROCESSING
# ==========================================
for share in shares:
    print(f"Fetching 1-min data for {share}...")

    folder_path = os.path.join(BASE_DIR, '1 min cash', str(cur_year), cur_month, cur_date)
    os.makedirs(folder_path, exist_ok=True)
    path = os.path.join(folder_path, f'{share} 1 min csh.xlsx')

    # Fetch 7 days of 1-minute data to cover weekends/holidays dynamically
    df_raw = yf.Ticker(share).history(period="7d", interval="1m")

    if df_raw.empty:
        print(f"No 1-min data found for {share}.")
        continue

    # Clean & normalize timestamps to US/Eastern
    df_raw.reset_index(inplace=True)
    time_col = 'Datetime' if 'Datetime' in df_raw.columns else 'Date'
    if df_raw[time_col].dt.tz is not None:
        df_raw[time_col] = df_raw[time_col].dt.tz_convert('US/Eastern').dt.tz_localize(None)

    df_raw['Date'] = df_raw[time_col].dt.date
    df_raw['Time_Str'] = df_raw[time_col].dt.strftime('%H:%M:%S')

    # Detect Trading Days dynamically
    unique_dates = sorted(df_raw['Date'].unique())

    if cur_date_datetime not in unique_dates:
        print(f"Target date {cur_date_datetime} not found in fetched data for {share}.")
        continue

    cur_idx = unique_dates.index(cur_date_datetime)
    prev_date_datetime = unique_dates[cur_idx - 1] if cur_idx > 0 else None

    # --- PROCESS PREVIOUS DAY DATA (>= 09:40 AM) ---
    if prev_date_datetime:
        df_prev = df_raw[(df_raw['Date'] == prev_date_datetime) &
                         (df_raw['Time_Str'] >= '09:40:00') &
                         (df_raw['Time_Str'] <= '16:00:00')]
        if not df_prev.empty:
            prev_high_dict[share] = round_to_005(df_prev['High'].max())
            prev_low_dict[share] = round_to_005(df_prev['Low'].min())
            prev_close_dict[share] = round_to_005(df_prev.iloc[-1]['Close'])

    # --- PROCESS TODAY'S DATA (>= 09:40 AM) ---
    df_today = df_raw[df_raw['Date'] == cur_date_datetime].copy()
    if df_today.empty:
        continue

    # Filter 1-min excel output to columns: Time, High, Low, Close, Volume
    df_export = df_today[['Time_Str', 'High', 'Low', 'Close', 'Volume']].copy()
    df_export.rename(columns={'Time_Str': 'Time'}, inplace=True)
    df_export = df_export.sort_values(by='Time')

    # Apply 0.05 rounding
    df_export['High'] = df_export['High'].apply(round_to_005)
    df_export['Low'] = df_export['Low'].apply(round_to_005)
    df_export['Close'] = df_export['Close'].apply(round_to_005)

    with pd.ExcelWriter(path) as writer:
        df_export.to_excel(writer, index=False)

    # Excel Formatting for 1-min file
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    # Convert string times to time objects
    for start_row in range(2, sheet.max_row + 1):
        time_cell = sheet.cell(start_row, 1)
        if isinstance(time_cell.value, str):
            try:
                time_val = datetime.datetime.strptime(time_cell.value, "%H:%M:%S")
                time_cell.value = time_val
                time_cell.number_format = 'h:mm AM/PM'
            except ValueError:
                pass
    wb.save(path)

    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    start_row = 2
    time_cell = sheet.cell(start_row, 1)
    target_time = datetime.datetime(1900, 1, 1, hour=9, minute=40)

    while time_cell.value is not None and time_cell.value < target_time:
        start_row += 1
        time_cell = sheet.cell(start_row, 1)

    start_row_2 = start_row
    end_time = datetime.datetime(1900, 1, 1, 16, 0, 0)
    cur_time = sheet.cell(start_row, 1).value

    # Store Today's 9:40 Close
    cl_9_40_dict[share] = round_to_005(sheet.cell(start_row, 4).value)

    # Highlight 9:40 row
    for i in range(1, 6):
        sheet.cell(start_row, i).fill = yellow_fill

    wb.save(path)
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    HIGH = 0
    LOW = 9999999
    temp_start_row = start_row

    # Today High / Low calculation
    while cur_time is not None and cur_time <= end_time and temp_start_row <= sheet.max_row:
        high_val = sheet.cell(temp_start_row, 2).value
        low_val = sheet.cell(temp_start_row, 3).value
        cur_time = sheet.cell(temp_start_row, 1).value

        if high_val is not None and high_val > HIGH:
            HIGH = high_val
        if low_val is not None and low_val < LOW and low_val != 0:
            LOW = low_val
        temp_start_row += 1

    high_dict[share] = round_to_005(HIGH)
    low_dict[share] = round_to_005(LOW)

    # Save 1-min aggregated 30-m stats
    sheet.cell(1, 7).value = "HIGH"
    sheet.cell(1, 8).value = "LOW"
    sheet.cell(1, 9).value = "CLOSE"

    start_row = start_row_2
    cur_time = sheet.cell(start_row, 1).value
    HIGH, LOW, count = 0, 9999999, 0

    while cur_time is not None and cur_time <= end_time and start_row <= sheet.max_row:
        high_val = sheet.cell(start_row, 2).value
        low_val = sheet.cell(start_row, 3).value

        if high_val is not None and high_val > HIGH:
            HIGH = high_val
        if low_val is not None and low_val < LOW and low_val != 0:
            LOW = low_val

        if count == 30:
            sheet.cell(start_row, 7).value = round_to_005(HIGH)
            sheet.cell(start_row, 8).value = round_to_005(LOW)

            close_val = sheet.cell(start_row, 4).value
            if close_val in [0, None]:
                temp_row = start_row
                while sheet.cell(temp_row, 4).value in [0, None] and temp_row > 1:
                    temp_row -= 1
                sheet.cell(start_row, 9).value = round_to_005(sheet.cell(temp_row, 4).value)
            else:
                sheet.cell(start_row, 9).value = round_to_005(close_val)

            count, HIGH, LOW = 1, 0, 9999999
            start_row += 1
            cur_time = sheet.cell(start_row, 1).value if start_row <= sheet.max_row else None
            continue

        start_row += 1
        count += 1
        cur_time = sheet.cell(start_row, 1).value if start_row <= sheet.max_row else None

    # Format time back to string
    for r in range(2, sheet.max_row + 1):
        time_cell = sheet.cell(r, 1)
        if isinstance(time_cell.value, datetime.datetime):
            time_cell.value = time_cell.value.strftime("%I:%M %p")
            time_cell.number_format = 'h:mm AM/PM'

    wb.save(path)
    print(f"1-min data complete for {share}")

# ==========================================
# 30-MINUTE (IMAGE 2 LAYOUT) & DAILY SUMMARY
# ==========================================
for share in shares:
    print(f"Processing Daily and 30-min data for {share}...")

    # --- 1. DAILY SUMMARY EXCEL ---
    daily_path = os.path.join(BASE_DIR, f'{share}.xlsx')

    if not os.path.exists(daily_path):
        raise FileNotFoundError(f"{share} daily file does not exist")

    start_daily = cur_date_datetime - datetime.timedelta(days=10)
    df_daily = yf.Ticker(share).history(interval='1d', start=start_daily.strftime('%Y-%m-%d'),
                                        end=(cur_date_datetime + datetime.timedelta(days=1)).strftime('%Y-%m-%d'))

    close = round_to_005(df_daily['Close'].iloc[-1]) if len(df_daily) >= 1 else 0
    vol = int(df_daily['Volume'].iloc[-1] // 100000) if len(df_daily) >= 1 else 0

    daily_wb = xl.load_workbook(daily_path)
    daily_sheet = daily_wb['D']
    daily_start_row = OFFSET

    daily_sheet.cell(daily_start_row, 1).value = f"{datetime.datetime.strptime(cur_date, "%d.%m.%y").strftime("%d/%m/%Y")}"
    daily_sheet.cell(daily_start_row, 1).alignment = alignment
    daily_sheet.cell(daily_start_row, 1).font = bold
    daily_sheet.cell(daily_start_row, 2).value = high_dict[share]
    daily_sheet.cell(daily_start_row, 2).font = blue
    daily_sheet.cell(daily_start_row, 2).alignment = alignment
    daily_sheet.cell(daily_start_row, 3).value = low_dict[share]
    daily_sheet.cell(daily_start_row, 3).font = red
    daily_sheet.cell(daily_start_row, 3).alignment = alignment
    daily_sheet.cell(daily_start_row, 4).value = close
    daily_sheet.cell(daily_start_row, 4).font = bold
    daily_sheet.cell(daily_start_row, 4).alignment = alignment
    daily_sheet.cell(daily_start_row, 5).value = vol
    daily_sheet.cell(daily_start_row, 5).font = bold
    daily_sheet.cell(daily_start_row, 5).alignment = alignment
    daily_sheet.cell(daily_start_row, 6).value = cl_9_40_dict[share]
    daily_sheet.cell(daily_start_row, 6).font = bold
    daily_sheet.cell(daily_start_row, 6).alignment = alignment
    daily_wb.save(daily_path)

    # --- 2. 30-MINUTE EXCEL (TARGET LAYOUT: IMAGE 2) ---
    folder_path_30m = os.path.join(BASE_DIR, '30 min cash', str(cur_year), cur_month, cur_date)
    os.makedirs(folder_path_30m, exist_ok=True)
    path = os.path.join(folder_path_30m, f'{share} 30 min csh.xlsx')

    df_raw_30 = yf.Ticker(share).history(interval="30m", start=cur_date_datetime.strftime('%Y-%m-%d'),
                                         end=(cur_date_datetime + datetime.timedelta(days=1)).strftime('%Y-%m-%d'))

    if df_raw_30.empty:
        continue

    df_raw_30.reset_index(inplace=True)
    time_col_30 = 'Datetime' if 'Datetime' in df_raw_30.columns else 'Date'

    if df_raw_30[time_col_30].dt.tz is not None:
        df_raw_30[time_col_30] = df_raw_30[time_col_30].dt.tz_convert('US/Eastern').dt.tz_localize(None)

    df_raw_30['Time'] = df_raw_30[time_col_30].dt.strftime('%I:%M %p')

    # Apply 0.05 rounding
    df_raw_30['High'] = df_raw_30['High'].apply(round_to_005)
    df_raw_30['Low'] = df_raw_30['Low'].apply(round_to_005)
    df_raw_30['Close'] = df_raw_30['Close'].apply(round_to_005)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = f'{share} Sheet'

    # --- ROW 1: HEADERS ---
    ws.cell(6, 6).value = share
    ws.cell(6, 7).value = "HIGH"
    ws.cell(6, 8).value = "LOW"
    ws.cell(6, 9).value = "LTP"

    # --- ROW 2: TODAY VALUES ---
    ws.cell(7, 6).value = cl_9_40_dict[share]
    ws.cell(7, 7).value = high_dict[share]
    ws.cell(7, 8).value = low_dict[share]
    ws.cell(7, 9).value = close

    # --- ROW 3: PREV DAY VALUES (YELLOW FILL) ---
    ws.cell(8, 6).value = "PREV"
    ws.cell(8, 7).value = prev_high_dict[share]
    ws.cell(8, 8).value = prev_low_dict[share]
    ws.cell(8, 9).value = prev_close_dict[share]

    for col in range(6, 10):
        ws.cell(8, col).fill = yellow_fill

    # --- ROW 4: TABLE HEADERS ---
    ws.cell(9, 6).value = "Time"
    ws.cell(9, 7).value = "High Rate"
    ws.cell(9, 8).value = "Low Rate"
    ws.cell(9, 9).value = "Close Rate"

    # --- ROW 5+: 30-MIN DATA ROWS ---
    for idx, row in df_raw_30.iterrows():
        r = idx + 10
        ws.cell(r, 6).value = row['Time']
        ws.cell(r, 7).value = row['High']
        ws.cell(r, 8).value = row['Low']
        ws.cell(r, 9).value = row['Close']

    # Apply styling across all populated cells
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=12):
        for cell in row:
            cell.font = bold
            cell.alignment = alignment

    # AUTOFIT 30-MIN SHEET
    autofit_columns(ws)

    wb.save(path)
    print(f"30-min data complete for {share}")

print("\nAll files successfully generated ")