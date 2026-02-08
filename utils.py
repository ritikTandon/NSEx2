from datetime import datetime
import shelve
from openpyxl.styles import Border, Side
from constants import DATE, bold, alignment, MAX_POINTS
import pandas as pd
import os
from urllib.parse import quote

def get_duration_params(from_datetime, to_datetime):
    """
    Convert input datetime strings to required query params format
    Example: 2026-01-23 09:15:00
    """

    from_date, from_time = from_datetime.split(" ")
    to_date, to_time = to_datetime.split(" ")

    from_datetime = f"{datetime.strptime(from_date, '%d.%m.%y').strftime('%Y-%m-%d')} {from_time}"
    to_datetime = f"{datetime.strptime(to_date, '%d.%m.%y').strftime('%Y-%m-%d')} {to_time}"

    return {
        "from": from_datetime,
        "to": to_datetime
    }

def get_fut_instrument_token(df: pd.DataFrame, is_expiry_today=False):
    """
    Method to get instrument token for nifty and bn based with
    additional check if expiry is today then get next expiry
    """
    # Ensure expiry is datetime
    df['expiry'] = pd.to_datetime(df['expiry'])

    result = {}

    for symbol in ["NIFTY", "BANKNIFTY"]:
        subset = df[
            (df["name"] == symbol) &
            (df["instrument_type"] == "FUT")
        ].sort_values("expiry")

        if subset.empty:
            continue

        # index 0 = front month, index 1 = next month
        row_index = 1 if is_expiry_today else 0

        # safety check
        if len(subset) > row_index:
            row = subset.iloc[row_index]
            result[row["tradingsymbol"]] = int(row["instrument_token"])

    return result


def add_missing_row(sheet, insert_row, date):
    """
    Add missing rows in scenarios where market is open on holidays
    """
    sheet.insert_rows(insert_row)   # row number where row will be inserted (this row will shift down)
    sheet.cell(insert_row, 1).value = date # Ex datetime(2026, 2, 1)
    sheet.cell(insert_row, 1).border = Border(right=Side(style='thin'))
    sheet.cell(insert_row, 1).number_format = "DD-MMM-YY"
    sheet.cell(insert_row, 1).font = bold
    sheet.cell(insert_row, 1).alignment = alignment


def sanitize_url(url: str) -> str:
    return quote(url, safe=":/?=#")