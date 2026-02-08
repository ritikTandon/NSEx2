from dotenv import load_dotenv
from kiteconnect import KiteConnect
import os, requests, logging, pandas as pd, io, json, datetime

from openpyxl.utils import get_column_letter

from constants import SYMBOL_DATA_API, HEADERS, EQ_SYMBOLS, DATE, FO_SYMBOLS, BASE_FOLDER_PATH, MONTH, YEAR, APPEND, \
    NO_FORMAT_LIST, FIXED_WIDTH, blue, alignment, red, bold, COPY_TO_CASH, SHARE_LIST, FO_SYMBOLS_WITH_EXPIRY, \
    LTP_DATA_API, LTP_PREV_PATH
from utils import get_duration_params
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl

load_dotenv()

api_key = os.getenv("API_KEY")
access_token = os.getenv("ACCESS_TOKEN")

kite = KiteConnect(api_key=api_key)
kite.set_access_token(access_token)


start_row = 9
start_col = 6  # Column F
idx = 2   # ltp prev sheet iterator

wb = openpyxl.load_workbook(LTP_PREV_PATH)
sheet_ltp_prev = wb["Sheet1"]

for symbol in SHARE_LIST:
    try:
        if symbol in EQ_SYMBOLS:
            INSTRUMENT_TOKEN = EQ_SYMBOLS[symbol][0]
        else:
            INSTRUMENT_TOKEN = FO_SYMBOLS[symbol][0]

        # 1 min
        URL = f"{SYMBOL_DATA_API}/{INSTRUMENT_TOKEN}/minute"
        PARAMS = get_duration_params(f"{DATE} 09:15:00", f"{DATE} 15:30:00")

        try:
            response = requests.get(URL, headers=HEADERS, params=PARAMS)
        except:
            print(f"Error in getting SYMBOL data for: {symbol}")
            idx += 1
            continue

        json_data = json.loads(response.text)

        data_list = json_data['data']['candles']

        # 2. Create DataFrame and Preprocess
        df = pd.DataFrame(data_list, columns=['Timestamp', 'Open', 'High', 'Low', 'Close', 'Volume'])

        df.drop(columns=['Open'], inplace=True)
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])
        df = df.sort_values('Timestamp')
        df.set_index('Timestamp', inplace=True)

        # ---------------------------------------------------------
        # CALCULATION 1: Day High and Low Vars (9:25 AM - 3:30 PM)
        # ---------------------------------------------------------
        start_time = datetime.time(9, 25)
        end_time = datetime.time(15, 30)

        # Create a mask for the specific time range
        day_mask = (df.index.time >= start_time) & (df.index.time <= end_time)
        day_df = df[day_mask]

        if not day_df.empty:
            day_high = day_df['High'].max()
            day_low = day_df['Low'].min()
        else:
            day_high = None
            day_low = None

        print(f"Day High (9:25-3:30): {day_high}")
        print(f"Day Low (9:25-3:30): {day_low}")

        # ---------------------------------------------------------
        # CALCULATION 2: 30-MIN AGGREGATES (TIME-BASED LOOP)
        # ---------------------------------------------------------

        df['30min_High'] = None
        df['30min_Low'] = None
        df['30min_Close'] = None

        calc_start = datetime.time(9, 25)
        calc_end = datetime.time(15, 30)

        HIGH = 0
        LOW = float('inf')

        block_start_time = datetime.time(9, 25)
        block_end_time = datetime.time(9, 55)  # first block special

        for ts in df.index:
            cur_time = ts.time()

            # Skip rows before 09:25 but KEEP them in df
            if cur_time < calc_start or cur_time > calc_end:
                continue

            high = df.at[ts, 'High']
            low = df.at[ts, 'Low']
            close = df.at[ts, 'Close']

            if pd.notna(high) and high > HIGH:
                HIGH = high

            if pd.notna(low) and low != 0 and low < LOW:
                LOW = low

            # Block end reached → write values
            if cur_time == block_end_time:
                df.at[ts, '30min_High'] = HIGH
                df.at[ts, '30min_Low'] = LOW
                df.at[ts, '30min_Close'] = close

                # Reset for next block
                HIGH = 0
                LOW = float('inf')

                # Move to next block
                next_start = (
                        datetime.datetime.combine(datetime.date.today(), block_end_time)
                        + datetime.timedelta(minutes=1)
                )

                next_end = next_start + datetime.timedelta(minutes=29)

                block_start_time = next_start.time()
                block_end_time = min(next_end.time(), calc_end)

        # Handle last partial block (15:26–15:30)
        last_ts = df.index[df.index.time <= calc_end][-1]
        if pd.isna(df.at[last_ts, '30min_High']):
            df.at[last_ts, '30min_High'] = HIGH
            df.at[last_ts, '30min_Low'] = LOW
            df.at[last_ts, '30min_Close'] = df.at[last_ts, 'Close']

        # ---------------------------------------------------------
        # EXPORT PREP
        # ---------------------------------------------------------
        df.reset_index(inplace=True)

        # Format Timestamp to "hh:mm am/pm" string
        df['Timestamp'] = df['Timestamp'].dt.strftime('%I:%M %p')

        # Rename column for final output
        df.rename(columns={'Timestamp': 'Time'}, inplace=True)

        # Define final column order
        export_cols = ['Time', 'High', 'Low', 'Close', 'Volume', '30min_High', '30min_Low', '30min_Close']
        df_final = df[export_cols]

        # ---------------------------------------------------------
        # OPENPYXL EXPORT & STYLING
        # ---------------------------------------------------------

        # Create a new Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = symbol

        # Write DataFrame to Worksheet
        for r in dataframe_to_rows(df_final, index=False, header=True):
            ws.append(r)

        # Define Yellow Fill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Iterate rows to find 9:25 AM and highlight 'Close'
        # Note: openpyxl is 1-indexed. Row 1 is header. Data starts at Row 2.
        # Columns: Time(A), Open(B), High(C), Low(D), Close(E)...

        target_time = "09:24 AM"
        time_col_idx = 1  # Column A

        # We iterate through the rows in the worksheet (skipping header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            time_cell = row[0]  # First cell in the row (Time)

            if time_cell.value == target_time:
                # Highlight the Close cell (index 4 in the row tuple, corresponding to column E)
                close_cell = row[3]
                close_cell.fill = yellow_fill
                close_925 = close_cell.value

        # Save the file
        if symbol not in FO_SYMBOLS:
            file_name = rf"{BASE_FOLDER_PATH}\1 MINUTE\{YEAR}\{MONTH}\{DATE}\{symbol}.xlsx"
            wb.save(file_name)
            print(f"{symbol} 1 MIN")

        # 15 and 30 min
        start_date = f"{(datetime.datetime.strptime(DATE, '%d.%m.%y') - datetime.timedelta(days=5)).strftime('%d.%m.%y')}"

        URL = f"{SYMBOL_DATA_API}/{INSTRUMENT_TOKEN}/day"
        PARAMS = get_duration_params(f"{start_date} 09:15:00", f"{DATE} 15:30:00")

        if symbol in FO_SYMBOLS:
            PARAMS |= {"continuous": "1"}   # in case the contract expired

        try:
            response = requests.get(URL, headers=HEADERS, params=PARAMS)
        except:
            print(f"Error in getting SYMBOL data for: {symbol}")
            idx += 1
            continue

        json_data = json.loads(response.text)

        data_list = json_data['data']['candles']

        # 3. VOL: Last item of the LAST candle divided by 100,000
        # // is floor division
        vol = data_list[-1][-1] // 100000

        # LTP AND PREV
        ltp = sheet_ltp_prev.cell(idx, 2).value
        prev = sheet_ltp_prev.cell(idx, 3).value

        # -----------------------------------
        # FETCH 15-MIN DATA
        # -----------------------------------
        URL = f"{SYMBOL_DATA_API}/{INSTRUMENT_TOKEN}/15minute"
        PARAMS = get_duration_params(
            f"{DATE} 09:15:00",
            f"{DATE} 15:30:00"
        )

        response = requests.get(URL, headers=HEADERS, params=PARAMS)
        json_data = json.loads(response.text)
        data_list = json_data["data"]["candles"]

        df_15 = pd.DataFrame(
            data_list,
            columns=["Timestamp", "Open", "High", "Low", "Close", "Volume"]
        )

        df_15.drop(columns=["Open", "Volume"], inplace=True)
        df_15["Timestamp"] = pd.to_datetime(df_15["Timestamp"])
        df_15 = df_15.sort_values("Timestamp").reset_index(drop=True)

        df_15['Timestamp'] = df_15['Timestamp'] + pd.Timedelta(minutes=15)

        # -----------------------------------
        # SAVE 15-MIN SHEET (NORMAL)
        # -----------------------------------
        df15_export = df_15.copy()
        df15_export["Time"] = df15_export["Timestamp"].dt.strftime("%I:%M %p")
        df15_final = df15_export[["Time", "High", "Low", "Close"]]

        wb15 = Workbook()
        ws15 = wb15.active
        ws15.title = symbol

        # HEADERS
        ws15.cell(6, 6).value = symbol
        ws15.cell(6, 7).value = "HIGH"
        ws15.cell(6, 8).value = "LOW"
        ws15.cell(6, 9).value = "LTP"
        ws15.cell(6, 10).value = "PREV"

        ws15.cell(8, 6).value = "Time"
        ws15.cell(8, 7).value = "High Rate"
        ws15.cell(8, 8).value = "Low Rate"
        ws15.cell(8, 9).value = "Close Rate"

        # FIXED VALUES
        ws15.cell(7, 6).value = close_925
        ws15.cell(7, 7).value = day_high
        ws15.cell(7, 8).value = day_low
        ws15.cell(7, 9).value = ltp
        ws15.cell(7, 10).value = prev

        # WRITE DATA (NO HEADERS)
        for r_idx, row in enumerate(
                dataframe_to_rows(df15_final, index=False, header=False)
        ):
            for c_idx, value in enumerate(row):
                ws15.cell(
                    row=start_row + r_idx,
                    column=start_col + c_idx,
                    value=value
                )

        # BOLD 50x50
        for row in ws15.iter_rows(min_row=1, max_row=50, min_col=1, max_col=50):
            for cell in row:
                cell.font = bold
                cell.alignment = alignment

        # increasing width
        for col_idx in range(6, 10):
            col_letter = get_column_letter(col_idx)
            ws15.column_dimensions[col_letter].width = FIXED_WIDTH

        file_15 = rf"{BASE_FOLDER_PATH}\15 MINUTE\{YEAR}\{MONTH}\{DATE}\{symbol}.xlsx"
        wb15.save(file_15)
        print(f"{symbol} 15 min")

        # -----------------------------------
        # BUILD CUSTOM 30-MIN FROM 15-MIN
        # -----------------------------------
        # First 15-min candle stays as-is
        rows = [{
            "Timestamp": df_15.iloc[0]["Timestamp"],
            "High": df_15.iloc[0]["High"],
            "Low": df_15.iloc[0]["Low"],
            "Close": df_15.iloc[0]["Close"]
        }]

        i = 1
        while i + 1 < len(df_15):
            first = df_15.iloc[i]
            second = df_15.iloc[i + 1]

            rows.append({
                "Timestamp": second["Timestamp"],
                "High": max(first["High"], second["High"]),
                "Low": min(first["Low"], second["Low"]),
                "Close": second["Close"]
            })

            i += 2

        df_30 = pd.DataFrame(rows)

        # -----------------------------------
        # SAVE CUSTOM 30-MIN SHEET
        # -----------------------------------
        df30_export = df_30.copy()
        df30_export["Time"] = pd.to_datetime(df30_export["Timestamp"]).dt.strftime("%I:%M %p")
        df30_final = df30_export[["Time", "High", "Low", "Close"]]

        wb30 = Workbook()
        ws30 = wb30.active
        ws30.title = symbol

        # HEADERS
        ws30.cell(6, 6).value = symbol
        ws30.cell(6, 7).value = "HIGH"
        ws30.cell(6, 8).value = "LOW"
        ws30.cell(6, 9).value = "LTP"
        ws30.cell(6, 10).value = "PREV"

        ws30.cell(8, 6).value = "Time"
        ws30.cell(8, 7).value = "High Rate"
        ws30.cell(8, 8).value = "Low Rate"
        ws30.cell(8, 9).value = "Close Rate"

        # FIXED VALUES
        ws30.cell(7, 6).value = close_925
        ws30.cell(7, 7).value = day_high
        ws30.cell(7, 8).value = day_low
        ws30.cell(7, 9).value = ltp
        ws30.cell(7, 10).value = prev

        # WRITE DATA
        for r_idx, row in enumerate(
                dataframe_to_rows(df30_final, index=False, header=False)
        ):
            for c_idx, value in enumerate(row):
                ws30.cell(
                    row=start_row + r_idx,
                    column=start_col + c_idx,
                    value=value
                )

        # BOLD 50x50
        for row in ws30.iter_rows(min_row=1, max_row=50, min_col=1, max_col=50):
            for cell in row:
                cell.font = bold
                cell.alignment = alignment

        # increasing width
        for col_idx in range(6, 10):
            col_letter = get_column_letter(col_idx)
            ws30.column_dimensions[col_letter].width = FIXED_WIDTH

        file_30 = rf"{BASE_FOLDER_PATH}\30 MINUTE\{YEAR}\{MONTH}\{DATE}\{symbol}.xlsx"
        wb30.save(file_30)

        if symbol in COPY_TO_CASH:
            file_30min_cash_raghav = rf"C:\Users\RITIK\Desktop\STUDY MATERIAL\hourlys 30 minute CASH\{YEAR}\{MONTH}\{DATE}\{symbol}.xlsx"
            wb30.save(file_30min_cash_raghav)

        print(f"{symbol} 30 min")

        # data feed
        if symbol not in FO_SYMBOLS:
            file_daily = rf"{BASE_FOLDER_PATH}\DAILY\{symbol}.xlsx"
            file_daily_cash = rf"{BASE_FOLDER_PATH}\CASH\{symbol}.xlsx"
            file_daily_raghav = rf"C:\Users\RITIK\Desktop\STUDY MATERIAL\CASH\DAILY\{symbol}.xlsx"
            file_daily_cash_raghav = rf"C:\Users\RITIK\Desktop\STUDY MATERIAL\CASH\{symbol}.xlsx"

            wb = openpyxl.load_workbook(file_daily)
            sheet = wb['D']

            input_row = EQ_SYMBOLS[symbol][1] + APPEND  # incrementing row values from base row(start row)

            # data filling
            sheet.cell(input_row, 2).value = day_high # high
            sheet.cell(input_row, 3).value = day_low # low
            sheet.cell(input_row, 4).value = ltp  # close
            sheet.cell(input_row, 5).value = ltp  # LTP
            sheet.cell(input_row, 6).value = vol  # vol
            sheet.cell(input_row, 7).value = close_925  # 9:25 close

            # number formatting
            if symbol not in NO_FORMAT_LIST:
                sheet.cell(input_row, 2).number_format = '0'  # high
                sheet.cell(input_row, 3).number_format = '0'  # low
                sheet.cell(input_row, 4).number_format = '0'  # close
                sheet.cell(input_row, 5).number_format = '0'  # LTP
                sheet.cell(input_row, 7).number_format = '0'  # 9:25 close

            # style formatting
            sheet.cell(input_row, 2).font = blue  # high
            sheet.cell(input_row, 2).alignment = alignment

            sheet.cell(input_row, 3).font = red  # low
            sheet.cell(input_row, 3).alignment = alignment

            sheet.cell(input_row, 4).font = bold  # close
            sheet.cell(input_row, 4).alignment = alignment

            sheet.cell(input_row, 5).font = bold  # LTP
            sheet.cell(input_row, 5).alignment = alignment

            sheet.cell(input_row, 6).font = bold  # vol
            sheet.cell(input_row, 6).alignment = alignment

            sheet.cell(input_row, 7).font = bold  # 9:25 close
            sheet.cell(input_row, 7).alignment = alignment

            # Activate sheet
            wb.active = wb.index(sheet)

            print(f"{symbol} done!")

            wb.save(file_daily)
            wb.save(file_daily_raghav)

            if symbol in COPY_TO_CASH:
                wb.save(file_daily_cash)
                wb.save(file_daily_cash_raghav)
        idx += 1
        print("<------------------------------------------------------>")

    except Exception as e:
        idx += 1
        print(f"\033[31mError occured for symbol: {symbol}\033[0m")
        print(f"\033[31mException: {e}\033[0m")