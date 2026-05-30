# from openpyxl.workbook import Workbook
# from datetime import datetime
# from constants import COPY_TO_CASH, BASE_FOLDER_PATH, EQ_SYMBOLS, APPEND
# import openpyxl as xl
#
# from new_share_create_update import get_last_row
import json

# d = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
#
# y = ["2026", "2025", "2024"]
#
# for a in y:
#     wb = Workbook()
#
#     for share in COPY_TO_CASH:
#         file_daily = rf"{BASE_FOLDER_PATH}\DAILY\{share}.xlsx"
#         close_wb = xl.load_workbook(file_daily)
#
#         cl_sheet = close_wb["Cl"]
#
#         close_row = get_last_row(cl_sheet)
#
#         close_date = (cl_sheet.cell(close_row, 1).value).split(" ")[-1].strftime("%d-%m-%Y")
#
#
#
#
#     wb.save(f"{a}.xlsx")

# for symbol in EQ_SYMBOLS:
#     file_daily = rf"{BASE_FOLDER_PATH}\DAILY\{symbol}.xlsx"
#     # print(file_daily)
#     wb = xl.load_workbook(file_daily)
#     s = wb['D']
#
#     input_row = EQ_SYMBOLS[symbol][1] + APPEND - 1
#
#     # print(s.cell(input_row, 1).value)
#     print(f"{s.cell(input_row, 2).value}\t{s.cell(input_row, 3).value}")
#
#     wb.close()


import openpyxl as xl
import requests

from constants import LTP_DATA_API, HEADERS, FO_SYMBOLS_WITH_EXPIRY, FO_SYMBOLS, EQ_SYMBOLS, LTP_PREV_PATH, \
    LTP_PREV_BACKUP_PATH, SYMBOL_DATA_API, DATE
from utils import sanitize_url, get_duration_params

# wb = xl.load_workbook(LTP_PREV_PATH)
# sheet = wb["Sheet1"]
#
# i = 2
# for symbol in {'BANKNIFTY': (16913410, 0)}:
#     # LTP AND PREV
#     instrument = f"NFO:{FO_SYMBOLS_WITH_EXPIRY[symbol]}"
#     URL = f"{LTP_DATA_API}{sanitize_url(f"{instrument}")}"
#
#     try:
#         response = requests.get(URL, headers=HEADERS)
#         print(response)
#     except:
#         print(f"Error in getting SYMBOL data for: {symbol}")
#         continue
#
#     try:
#         ltp = response.json()["data"][instrument]["last_price"]
#     except:
#         print(f"LTP not found for {symbol}")
#         continue
#
#     print(ltp)

for symbol in {'BANKNIFTY': (16913410, 0)}:
    URL = f"{SYMBOL_DATA_API}/16914178/day"
    PARAMS = get_duration_params(f"{DATE} 09:15:00", f"{DATE} 15:30:00")

    try:
        response = requests.get(URL, headers=HEADERS, params=PARAMS)
    except:
        print(f"Error in getting SYMBOL data for: {symbol}")
        continue

    json_data = json.loads(response.text)
    print(json_data)
