import openpyxl as xl
import requests

from constants import LTP_DATA_API, HEADERS, FO_SYMBOLS_WITH_EXPIRY, FO_SYMBOLS, EQ_SYMBOLS, LTP_PREV_PATH
from utils import sanitize_url

wb = xl.load_workbook(LTP_PREV_PATH)
sheet = wb["Sheet1"]

i = 2
for symbol in EQ_SYMBOLS | FO_SYMBOLS:
    # LTP AND PREV
    if symbol in FO_SYMBOLS:
        instrument = f"NFO:{FO_SYMBOLS_WITH_EXPIRY[symbol]}"
        URL = f"{LTP_DATA_API}{sanitize_url(f"{instrument}")}"

        try:
            response = requests.get(URL, headers=HEADERS)
        except:
            print(f"Error in getting SYMBOL data for: {symbol}")
            continue

        try:
            ltp = response.json()["data"][instrument]["last_price"]
        except:
            print(f"LTP not found for {symbol}")
            continue

    else:
        instrument = f"NSE:{symbol}"
        URL = f"{LTP_DATA_API}{sanitize_url(f"{instrument}")}"
        try:
            response = requests.get(URL, headers=HEADERS)
        except:
            print(f"Error in getting SYMBOL data for: {symbol}")
            continue

        try:
            ltp = response.json()["data"][instrument]["last_price"]
        except:
            print(f"LTP not found for {symbol}")
            continue

    prev = sheet.cell(i, 2).value

    sheet.cell(i, 2).value = ltp
    sheet.cell(i, 3).value = prev

    i += 1

    wb.save(LTP_PREV_PATH)